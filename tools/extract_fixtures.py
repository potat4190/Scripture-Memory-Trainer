"""Regenerate seed/cards.json and tests/fixtures/*.json from the challenge workbook.

The workbook is the single source of truth for all three files. Strings are
copied out of the cells verbatim -- no Unicode normalization, no whitespace
trimming -- because the exact codepoints (and, for one case, the exact leading
and trailing spaces) are what the CHECK cases test.

The workbook is not tracked in this repo -- the generated JSON files are, so the
repo stands on its own and this script is only needed to re-derive them. Put your
copy of the workbook at the path below before running.

openpyxl is only needed for this one-shot extraction, so it is not a project
dependency. Run:

    uv run --with openpyxl python tools/extract_fixtures.py

Then re-run the test suite; tests/test_fixtures.py asserts the shape of what
this writes.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "docs" / "Frontier Commons Fellowship FA26 — Build Lane Challenge Data.xlsx"

# The B_* sheets carry two title rows and a blank spacer above the header.
CARDS_HEADER_ROW = 1
CHECK_FIRST_DATA_ROW = 6

# B_Check_Answers spells an empty input this way rather than leaving the cell blank.
EMPTY_INPUT_SENTINEL = "(empty string)"


def _write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {len(payload):>2} rows -> {path.relative_to(ROOT)}")


def extract_cards(wb: Any) -> list[dict[str, str]]:
    ws = wb["B_Cards"]
    header = [c.value for c in ws[CARDS_HEADER_ROW]]
    cards: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=CARDS_HEADER_ROW + 1, max_row=ws.max_row):
        values = [c.value for c in row]
        if values[0] is None:
            continue
        cards.append(dict(zip(header, values, strict=True)))
    return cards


def extract_schedule(wb: Any) -> list[dict[str, Any]]:
    ws = wb["B_Check_Schedule"]
    traces: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=CHECK_FIRST_DATA_ROW, max_row=ws.max_row):
        _, num, grades, trace, end_box, next_due = (c.value for c in row)
        if num is None:
            continue
        traces.append(
            {
                "id": int(num),
                # B_Check_Schedule's preamble: "Every card starts at box 0 on 2026-09-01."
                "start_box": 0,
                "start_date": "2026-09-01",
                "grades": [g.strip() for g in grades.split(",")],
                "trace": trace,
                "expected_box": int(end_box),
                "expected_due": next_due,
            }
        )
    return traces


def extract_answers(wb: Any) -> list[dict[str, Any]]:
    ws = wb["B_Check_Answers"]
    cases: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=CHECK_FIRST_DATA_ROW, max_row=ws.max_row):
        _, reference, lang, change, raw_input, verdict = (c.value for c in row)
        if reference is None:
            continue
        cases.append(
            {
                "id": len(cases) + 1,
                "reference": reference,
                "lang": lang,
                "change": change,
                "input": "" if raw_input == EMPTY_INPUT_SENTINEL else raw_input,
                "expected_verdict": verdict,
            }
        )
    return cases


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(
            f"Workbook not found: {WORKBOOK} -- it is not tracked in this repo. "
            "Place your copy there and re-run."
        )
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    _write(ROOT / "seed" / "cards.json", extract_cards(wb))
    _write(ROOT / "tests" / "fixtures" / "check_schedule.json", extract_schedule(wb))
    _write(ROOT / "tests" / "fixtures" / "check_answers.json", extract_answers(wb))


if __name__ == "__main__":
    main()
